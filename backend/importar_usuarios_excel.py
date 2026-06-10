"""
Script de importacion de usuarios reales desde el archivo Excel clients.xlsx.
Mapea las columnas del Excel a los campos del modelo Deportista de BJJBFIT.

Columnas del Excel:
  1-Nombre, 2-Apellidos, 3-Telefono, 4-Categoria, 5-Sexo, 6-Correo,
  7-Direccion, 8-Localidad, 9-Provincia, 10-Codigo postal, 11-Pais,
  12-DNI, 13-Reservas finalizadas, 14-Numero de cuenta,
  15-Notas, 16-Fecha de Nacimiento, 17-Como nos has conocido?,
  18-Pagos recurrentes
"""
import os
import sys
import re
import unicodedata
import django

# Configuracion Django
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

import openpyxl
from datetime import date
from django.contrib.auth.hashers import make_password
from usuarios.models import Deportista, Plan


# =====================================================================
# CONSTANTES
# =====================================================================
EXCEL_PATH = r'C:\Users\Victor\Downloads\clients.xlsx'
CONTRASENA_TEMPORAL = 'BjjbFit2026!'

# Filas a descartar (basura/pruebas detectadas en el analisis)
FILAS_DESCARTADAS = {2, 215, 339, 340}  # Fila 2: "-/-", 215: John Doe, 339/340: Prueba 01/03

# Diccionario de nombres espanoles para inferir sexo
# 'M' = Masculino, 'F' = Femenino
NOMBRES_MASCULINOS = {
    'aaron', 'abel', 'abraham', 'abubakr', 'adria', 'adrian',
    'aitor', 'alan', 'albert', 'aleix', 'alejandro', 'alex',
    'alfredo', 'alvaro', 'amine', 'andres', 'angel',
    'antonio', 'arnau', 'arturo', 'asier',
    'borja', 'brahim', 'bruno',
    'carlos', 'cesar', 'christian', 'cristian',
    'dani', 'daniel', 'daoud', 'david', 'deiby', 'diego', 'dmytro',
    'edgar', 'edilson', 'eduardo', 'eloy', 'emilio', 'enrique', 'eric', 'erik',
    'ernesto', 'esteban', 'eyal',
    'fco', 'felipe', 'fernando', 'fran', 'francisco',
    'gabriel', 'gerard', 'german', 'gonzalo', 'guillem', 'guillermo',
    'gustavo',
    'habil', 'hector', 'hicham', 'hugo',
    'ignacio', 'ivan',
    'jaime', 'jaume', 'javi', 'javier', 'jean', 'jesus', 'joan',
    'joel', 'john', 'jordi', 'jorge', 'jose', 'josep', 'juan',
    'juanjo', 'julio',
    'kevin',
    'leo', 'leonardo', 'lluis', 'lucas', 'luis',
    'manu', 'manuel', 'marc', 'marcos', 'mario', 'martin',
    'mateo', 'matias', 'miguel', 'miquel', 'moises',
    'neo', 'nestor', 'nico', 'nicolas',
    'omar', 'oriol', 'oscar',
    'pablo', 'pau', 'pedro', 'pol',
    'quim',
    'rafa', 'rafael', 'ramon', 'raul', 'ricardo',
    'robert', 'rodney', 'rodrigo', 'roger', 'ruben',
    'salvador', 'samuel', 'santi', 'santiago', 'saul',
    'sebastian', 'sergi', 'sergio',
    'thiago', 'tomas',
    'valentin', 'victor', 'vicente',
    'xavi', 'xavier',
    'yaddiel', 'yeray', 'yovanny',
}

NOMBRES_FEMENINOS = {
    'aina', 'alba', 'alicia', 'ana', 'andrea', 'angela',
    'ariadna', 'aurora',
    'beatriz', 'berta',
    'carla', 'carmen', 'carolina', 'catalina', 'cecilia', 'clara',
    'claudia', 'cristina',
    'diana',
    'elena', 'elisabet', 'elisabeth', 'elisa', 'emma', 'esther', 'eva',
    'gloria',
    'helena',
    'ines', 'irene', 'isabel', 'isabella',
    'jessica', 'joana', 'julia', 'julieta',
    'laura', 'laia', 'leire', 'lidia', 'lorena', 'lucia',
    'luisa',
    'manuela', 'mar', 'marga', 'maria', 'marina', 'marta',
    'martina', 'mercedes', 'mireia', 'miriam', 'montse', 'montserrat',
    'naiara', 'nayara', 'nerea', 'noelia', 'noemi',
    'nuria',
    'olga',
    'patricia', 'paula', 'pilar',
    'raquel', 'rebeca', 'rocio', 'rosa', 'ruth',
    'saida', 'sandra', 'sara', 'silvia', 'sonia', 'sofia',
    'susana',
    'teresa',
    'valentina', 'vanessa', 'veronica', 'virginia',
    'yolanda',
    'zoe',
}


# =====================================================================
# FUNCIONES AUXILIARES
# =====================================================================

def normalizar_texto(texto: str) -> str:
    """Elimina acentos y convierte a minusculas para comparaciones."""
    if not texto:
        return ''
    nfkd = unicodedata.normalize('NFKD', texto)
    sin_acentos = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return sin_acentos.lower().strip()


def generar_username(nombre: str, apellidos: str, usernames_existentes: set) -> str:
    """
    Genera un username unico a partir del nombre y apellidos.
    Formato: nombre.apellido1 (en minusculas, sin acentos, sin espacios).
    Si hay duplicados, anade un sufijo numerico.
    """
    nombre_norm = normalizar_texto(nombre).split()[0] if nombre else 'usuario'
    apellido_parts = normalizar_texto(apellidos).split() if apellidos else ['sin_apellido']
    apellido_norm = apellido_parts[0] if apellido_parts else 'sin_apellido'
    
    # Limpiar caracteres no alfanumericos
    nombre_norm = re.sub(r'[^a-z0-9]', '', nombre_norm)
    apellido_norm = re.sub(r'[^a-z0-9]', '', apellido_norm)
    
    if not nombre_norm:
        nombre_norm = 'usuario'
    if not apellido_norm:
        apellido_norm = 'x'
    
    base = f"{nombre_norm}.{apellido_norm}"
    username = base
    contador = 1
    while username in usernames_existentes:
        username = f"{base}{contador}"
        contador += 1
    
    usernames_existentes.add(username)
    return username


def inferir_sexo(nombre: str) -> str | None:
    """
    Infiere el sexo ('M' o 'F') a partir del primer nombre.
    Devuelve None si no se puede determinar.
    """
    if not nombre or nombre == '-':
        return None
    
    # Tomar solo el primer nombre
    primer_nombre = nombre.strip().split()[0].strip()
    primer_nombre_norm = normalizar_texto(primer_nombre)
    
    # Buscar en ambos diccionarios (normalizado)
    if primer_nombre_norm in NOMBRES_MASCULINOS:
        return 'M'
    if primer_nombre_norm in NOMBRES_FEMENINOS:
        return 'F'
    
    return None


def determinar_categoria_edad(fecha_nacimiento) -> str:
    """Determina la categoria de edad segun la fecha de nacimiento."""
    if not fecha_nacimiento:
        return 'ADULTO'  # Por defecto si no hay fecha
    
    if hasattr(fecha_nacimiento, 'date'):
        fecha = fecha_nacimiento.date()
    else:
        fecha = fecha_nacimiento
    
    hoy = date.today()
    edad = (hoy - fecha).days // 365
    
    if edad <= 9:
        return 'INFANTIL'
    elif edad <= 17:
        return 'JUVENIL'
    else:
        return 'ADULTO'


def limpiar_telefono(telefono) -> str:
    """Normaliza el telefono a un string limpio."""
    if not telefono:
        return ''
    tel = str(telefono).strip()
    # Quitar posibles decimales (.0 de Excel)
    if tel.endswith('.0'):
        tel = tel[:-2]
    # Quitar espacios y guiones
    tel = re.sub(r'[\s\-\.\(\)]', '', tel)
    return tel


def limpiar_dni(dni) -> str:
    """Normaliza el DNI/NIF."""
    if not dni:
        return ''
    return str(dni).strip().upper().replace(' ', '')


def normalizar_cuenta_bancaria(cuenta) -> str:
    """Limpia y normaliza la cuenta bancaria (IBAN)."""
    if not cuenta:
        return ''
    return str(cuenta).strip().upper().replace(' ', '').replace('-', '')


# =====================================================================
# FUNCION PRINCIPAL DE IMPORTACION
# =====================================================================

def importar(ruta_excel: str = EXCEL_PATH):
    """Lee el Excel y crea los usuarios en la base de datos."""
    
    print("=" * 60)
    print("  IMPORTACION DE USUARIOS - BJJBFIT")
    print("=" * 60)
    
    # Cargar Excel
    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb.active
    total_filas = ws.max_row - 1  # Sin cabecera
    print(f"\n[INFO] Archivo: {ruta_excel}")
    print(f"[INFO] Total filas en Excel: {total_filas}")
    print(f"[INFO] Filas descartadas (basura): {len(FILAS_DESCARTADAS)}")
    
    # Obtener planes existentes
    planes = {}
    for plan in Plan.objects.all():
        cat = plan.categoria_edad
        # Preferir los planes genericos (no fundador) como default
        if cat not in planes or 'fundador' not in plan.nombre.lower():
            planes[cat] = plan
    print(f"\n[PLANES] Planes mapeados:")
    for cat, plan in planes.items():
        print(f"   {cat} -> {plan.nombre} ({plan.precio_base} EUR)")
    
    # Recolectar usernames existentes
    usernames_existentes = set(Deportista.objects.values_list('username', flat=True))
    emails_existentes = set(
        e.lower() for e in Deportista.objects.values_list('email', flat=True) if e
    )
    
    # Contadores para el informe
    contadores = {
        'importados': 0,
        'descartados': 0,
        'sexo_inferido': 0,
        'sexo_excel': 0,
        'sexo_desconocido': 0,
        'sin_email_placeholder': 0,
        'con_cuenta_bancaria': 0,
        'duplicados_email': 0,
        'errores': [],
        'relaciones_padre_hijo': [],
    }
    
    # Primera pasada: recolectar DNIs para detectar relaciones padre-hijo
    dnis_map = {}  # dni -> [(fila, nombre, apellidos, fecha_nac)]
    for row in range(2, ws.max_row + 1):
        if row in FILAS_DESCARTADAS:
            continue
        dni = ws.cell(row=row, column=12).value
        if dni:
            dni_clean = limpiar_dni(dni)
            if dni_clean:
                nombre = ws.cell(row=row, column=1).value
                apellidos = ws.cell(row=row, column=2).value
                fecha = ws.cell(row=row, column=16).value
                if dni_clean not in dnis_map:
                    dnis_map[dni_clean] = []
                dnis_map[dni_clean].append((row, nombre, apellidos, fecha))
    
    # Identificar familias (DNI compartido donde hay adulto + menor)
    familias = {k: v for k, v in dnis_map.items() if len(v) > 1}
    
    # Mapeo de fila -> usuario creado (para asignar padre_tutor despues)
    fila_a_deportista = {}
    # Mapeo de fila -> fila_padre
    fila_padre_map = {}
    # Set de filas que son duplicados reales (mismo nombre, mismo DNI) para descartar
    filas_duplicadas = set()
    
    # Identificar en familias quien es padre y quien hijo, y duplicados
    for dni, miembros in familias.items():
        # Detectar duplicados reales (mismo nombre normalizado)
        nombres_norm = {}
        for fila, nombre, apellidos, fecha in miembros:
            key = normalizar_texto(f"{nombre} {apellidos}")
            if key not in nombres_norm:
                nombres_norm[key] = []
            nombres_norm[key].append((fila, fecha))
        
        # Marcar duplicados (conservar el que tiene fecha de nacimiento, o el primero)
        for key, filas_grupo in nombres_norm.items():
            if len(filas_grupo) > 1:
                # Ordenar: primero los que tienen fecha, luego el primero
                con_fecha = [(f, d) for f, d in filas_grupo if d]
                sin_fecha = [(f, d) for f, d in filas_grupo if not d]
                conservar = con_fecha[0][0] if con_fecha else filas_grupo[0][0]
                for fila, _ in filas_grupo:
                    if fila != conservar:
                        filas_duplicadas.add(fila)
        
        # Ahora identificar padre-hijo (miembros no duplicados)
        miembros_validos = [(f, n, a, d) for f, n, a, d in miembros if f not in filas_duplicadas]
        adultos = []
        menores = []
        for fila, nombre, apellidos, fecha in miembros_validos:
            if fecha:
                if hasattr(fecha, 'date'):
                    f = fecha.date()
                else:
                    f = fecha
                edad = (date.today() - f).days // 365
                if edad >= 18:
                    adultos.append(fila)
                else:
                    menores.append(fila)
            else:
                adultos.append(fila)
        
        if adultos and menores:
            padre_fila = adultos[0]
            for hijo_fila in menores:
                fila_padre_map[hijo_fila] = padre_fila
                contadores['relaciones_padre_hijo'].append(
                    (hijo_fila, padre_fila, dni)
                )
    
    print(f"\n[INFO] Duplicados detectados por nombre+DNI: {len(filas_duplicadas)}")
    
    # Hash de la contrasena temporal (una sola vez para eficiencia)
    password_hash = make_password(CONTRASENA_TEMPORAL)
    
    # Segunda pasada: crear usuarios
    print(f"\n[IMPORT] Importando usuarios...")
    
    for row in range(2, ws.max_row + 1):
        if row in FILAS_DESCARTADAS or row in filas_duplicadas:
            contadores['descartados'] += 1
            continue
        
        # Leer campos del Excel
        nombre = ws.cell(row=row, column=1).value
        apellidos = ws.cell(row=row, column=2).value
        telefono = ws.cell(row=row, column=3).value
        sexo_excel = ws.cell(row=row, column=5).value
        correo = ws.cell(row=row, column=6).value
        dni = ws.cell(row=row, column=12).value
        cuenta = ws.cell(row=row, column=14).value
        fecha_nac = ws.cell(row=row, column=16).value
        
        # Limpiar nombre
        nombre = str(nombre).strip() if nombre and nombre != '-' else ''
        apellidos = str(apellidos).strip() if apellidos and apellidos != '-' else ''
        
        if not nombre and not apellidos:
            contadores['descartados'] += 1
            continue
        
        # Generar username
        username = generar_username(nombre, apellidos, usernames_existentes)
        
        # Resolver email
        email = str(correo).strip().lower() if correo else ''
        if not email:
            email = f"{username}@bjjbfit-placeholder.com"
            contadores['sin_email_placeholder'] += 1
        elif email in emails_existentes:
            base_parts = email.split('@')
            email = f"{base_parts[0]}+{row}@{base_parts[1]}"
            contadores['duplicados_email'] += 1
        emails_existentes.add(email)
        
        # Resolver sexo
        # En el Excel: 'H' = Hombre, 'M' = Mujer
        # En nuestro modelo: 'M' = Masculino, 'F' = Femenino
        if sexo_excel:
            sexo_str = str(sexo_excel).strip().upper()
            if sexo_str == 'H':
                sexo = 'M'  # Hombre -> Masculino
            elif sexo_str == 'M':
                sexo = 'F'  # Mujer -> Femenino (OJO: en el Excel 'M' = Mujer)
            elif sexo_str == 'F':
                sexo = 'F'
            else:
                sexo = inferir_sexo(nombre)
            contadores['sexo_excel'] += 1
        else:
            sexo = inferir_sexo(nombre)
            if sexo:
                contadores['sexo_inferido'] += 1
            else:
                contadores['sexo_desconocido'] += 1
        
        # Resolver fecha de nacimiento
        fecha_nacimiento = None
        if fecha_nac:
            if hasattr(fecha_nac, 'date'):
                fecha_nacimiento = fecha_nac.date()
            elif hasattr(fecha_nac, 'year'):
                fecha_nacimiento = fecha_nac
        
        # Determinar categoria de edad
        categoria = determinar_categoria_edad(fecha_nacimiento)
        
        # Asignar plan segun categoria (sin activar)
        tipo_plan = planes.get(categoria)
        
        # Cuenta bancaria
        cuenta_limpia = normalizar_cuenta_bancaria(cuenta)
        metodo_pago = 'CUENTA' if cuenta_limpia else 'EFECTIVO'
        if cuenta_limpia:
            contadores['con_cuenta_bancaria'] += 1
        
        # Crear el deportista
        try:
            deportista = Deportista(
                username=username,
                email=email,
                password=password_hash,
                first_name=nombre,
                last_name=apellidos,
                telefono=limpiar_telefono(telefono),
                nif=limpiar_dni(dni),
                fecha_nacimiento=fecha_nacimiento,
                sexo=sexo,
                cinturon='Blanco',
                grados=0,
                cuenta_bancaria=cuenta_limpia if cuenta_limpia else None,
                metodo_pago=metodo_pago,
                plan_activo=False,  # Admin los activa manualmente
                tipo_plan=tipo_plan,
                es_familiar=False,
                requiere_cambio_password=True,
                is_active=True,
            )
            deportista.save()
            fila_a_deportista[row] = deportista
            contadores['importados'] += 1
            
            if contadores['importados'] % 50 == 0:
                print(f"   ... {contadores['importados']} usuarios importados")
                
        except Exception as e:
            contadores['errores'].append((row, nombre, apellidos, str(e)))
            print(f"   [ERROR] Fila {row} ({nombre} {apellidos}): {e}")
    
    # Tercera pasada: asignar relaciones padre-hijo
    print(f"\n[FAMILIA] Asignando relaciones padre-hijo...")
    relaciones_asignadas = 0
    for hijo_fila, padre_fila, dni in contadores['relaciones_padre_hijo']:
        padre = fila_a_deportista.get(padre_fila)
        hijo = fila_a_deportista.get(hijo_fila)
        if padre and hijo:
            hijo.padre_tutor = padre
            hijo.save(update_fields=['padre_tutor'])
            relaciones_asignadas += 1
            print(f"   [OK] {hijo.first_name} {hijo.last_name} -> padre: {padre.first_name} {padre.last_name} (DNI: {dni})")
    
    # =====================================================================
    # INFORME FINAL
    # =====================================================================
    print("\n" + "=" * 60)
    print("  INFORME DE IMPORTACION")
    print("=" * 60)
    print(f"  Total filas en Excel:     {total_filas}")
    print(f"  Usuarios importados:      {contadores['importados']}")
    print(f"  Descartados (basura/dup): {contadores['descartados']}")
    print(f"  Errores:                  {len(contadores['errores'])}")
    print(f"  ---")
    print(f"  Sexo del Excel:           {contadores['sexo_excel']}")
    print(f"  Sexo inferido por nombre: {contadores['sexo_inferido']}")
    print(f"  Sexo desconocido (null):  {contadores['sexo_desconocido']}")
    print(f"  ---")
    print(f"  Sin email (placeholder):  {contadores['sin_email_placeholder']}")
    print(f"  Emails duplicados:        {contadores['duplicados_email']}")
    print(f"  Con cuenta bancaria:      {contadores['con_cuenta_bancaria']}")
    print(f"  Relaciones padre-hijo:    {relaciones_asignadas}")
    print(f"  ---")
    print(f"  Contrasena temporal:      {CONTRASENA_TEMPORAL}")
    print(f"  Cambio obligatorio:       Si (requiere_cambio_password=True)")
    print(f"  Plan activo:              No (plan_activo=False, admin lo activa)")
    
    if contadores['errores']:
        print(f"\n  ERRORES DETALLADOS:")
        for fila, nombre, apellidos, error in contadores['errores']:
            print(f"    Fila {fila}: {nombre} {apellidos} -> {error}")
    
    # Verificacion final en BD
    total_bd = Deportista.objects.filter(is_staff=False).count()
    print(f"\n  Total deportistas en BD (sin admin): {total_bd}")
    
    print("\n[OK] Importacion completada.")
    return contadores


if __name__ == '__main__':
    importar()
